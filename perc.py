import streamlit as st
import pandas as pd
import io
from collections import OrderedDict
import re
import numpy as np
 
# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Convertir TXT Perceptron a Excel", layout="wide")
 
# --- ESTILO GLOBAL (FONDO OSCURO, TABLAS CLARAS) ---
st.markdown("""
    <style>
    body {
        background-color: #121212;
        color: #FFFFFF;
        font-family: 'Poppins', sans-serif;
    }
 
    .stApp {
        background-color: #121212;
    }
 
    /* Encabezados */
    h1, h2, h3, h4 {
        color: #ffc107;
    }
 
    /* Área de subida de archivos */
    div[data-testid="stFileUploader"] {
        border: 2px dashed #5a5a5a !important;
        background-color: rgba(50,50,50,0.7);
        border-radius: 15px;
        padding: 20px;
    }
 
    div[data-testid="stFileUploader"]:hover {
        border-color: #ffc107 !important;
        background-color: rgba(80,80,80,0.9);
    }
 
    /* Tabla de correlación */
    .dataframe {
        background: #2b2b2b !important;
        color: #ffffff !important;
        border-radius: 10px;
        font-size: 15px;
    }
 
    .dataframe td, .dataframe th {
        text-align: center !important;
        padding: 8px !important;
    }
 
    /* Botón de descarga */
    div.stDownloadButton > button {
        background-color: #ffc107;
        color: #000;
        font-weight: bold;
        border-radius: 10px;
        border: none;
        padding: 10px 25px;
    }
 
    div.stDownloadButton > button:hover {
        background-color: #ffde59;
        color: #000;
    }
    </style>
""", unsafe_allow_html=True)
 
# --- TÍTULO ---
st.title("📄 Comparativo Frontal vs Final")
 
# --- FUNCIÓN PARA PROCESAR TXT ---
def procesar_txt_a_df(archivo):
    contenido = archivo.read().decode("latin-1").splitlines()
    encabezados, mediciones = [], []
 
    for linea in contenido:
        partes = linea.strip().split("\t")
        if "JSN" in partes and "PSN" in partes:
            encabezados = partes
            continue
        if partes and partes[0].upper() not in ["NOMINAL", "USL", "LSL", "UTL", "LTL", "URL", "LRL"]:
            mediciones.append(partes)
 
    if not encabezados or not mediciones:
        return None, []
 
    filas_med = []
    for med in mediciones:
        fila = OrderedDict({
            "JSN": med[0],
            "PSN": med[1] if len(med) > 1 else "",
            "Fecha": med[2] if len(med) > 2 else "",
            "Hora": med[3] if len(med) > 3 else ""
        })
        for i, col in enumerate(encabezados[4:], start=4):
            fila[col] = med[i] if i < len(med) else ""
        filas_med.append(fila)
 
    eje_cols = encabezados[4:]
    return pd.DataFrame(filas_med), eje_cols
 

# --- MAPEO DE EJES ---
def map_axis(front_axis):

    forced_map = {
        "1000L[Y]": "1000L[Y]",
        "1000R[Y]": "1000R[Y]",
        "1100L[X]": "3125L[X]",
        "1100L[Y]": "3125L[Y]",
        "1100L[Z]": "3125L[Z]",
        "1100R[X]": "3125R[X]",
        "1100R[Y]": "3125R[Y]",
        "1100R[Z]": "3125R[Z]"
    }

    if front_axis in forced_map:
        return forced_map[front_axis]

    match = re.match(r"(1100)([LR]\[[XYZ]\])", front_axis)
    if match:
        return f"3125{match.group(2)}"

    return front_axis

# --- SUBIDA DE ARCHIVOS ---
st.subheader("📤 Archivo Frontal/Trasero")
archivo_frontal = st.file_uploader("Carga el archivo TXT Frontal de Perceptron", type=["txt"], key="frontal")
 
st.subheader("📤 Archivo Final")
archivo_final = st.file_uploader("Carga el archivo TXT Final de Perceptron", type=["txt"], key="final")
 
# --- PROCESAMIENTO ---
if archivo_frontal and archivo_final:
    df_frontal, ejes_frontal = procesar_txt_a_df(archivo_frontal)
    df_final, ejes_final = procesar_txt_a_df(archivo_final)
 
    if df_frontal is None or df_final is None:
        st.error("⚠️ Uno de los archivos no contiene mediciones reales válidas.")
    else:
        st.success("✅ Archivos procesados correctamente. Descarga combinada lista.")
 
        df_frontal["PSN"] = df_frontal["PSN"].astype(str).str.strip()
        df_final["PSN"] = df_final["PSN"].astype(str).str.strip()

        df_match = pd.DataFrame({
            "PSN": list(
                set(df_frontal["PSN"]).intersection(
                    set(df_final["PSN"])
                )
            )
        }).sort_values("PSN").reset_index(drop=True)

        psn_validos = df_match["PSN"].unique()

        df_frontal = df_frontal[df_frontal["PSN"].isin(psn_validos)].reset_index(drop=True)
        df_final = df_final[df_final["PSN"].isin(psn_validos)].reset_index(drop=True)

        # ---------------------------------
        # EXTRAER STATION Y MODEL DINÁMICO
        # ---------------------------------

        nombre_archivo = archivo_frontal.name.replace(".txt", "")

        # Quitar timestamp inicial (todo antes del primer "_")
        if "_" in nombre_archivo:
            nombre_sin_fecha = nombre_archivo.split("_", 1)[1]
        else:
            nombre_sin_fecha = nombre_archivo

        # Separar partes
        partes = nombre_sin_fecha.split("_")

        if "Front" in partes:
            idx_front = partes.index("Front")

            # Station → todo hasta "Mod"
            station_name = "_".join(partes[:idx_front+2])

            # Modelo → lo que viene después
            model_name = "_".join(partes[idx_front+2:]) if len(partes) > idx_front+2 else "UNKNOWN"
        else:
            station_name = nombre_sin_fecha
            model_name = "UNKNOWN"
        
        # ---------------------------
        # 🔥 GENERAR MAPEOS DE EJES
        # ---------------------------
        ejes_mapeados = []
        for eje in ejes_frontal:
            eje_final = map_axis(eje)
            if eje_final in ejes_final or eje_final.startswith("3125"):
                ejes_mapeados.append((eje, eje_final))

        df_axes = pd.DataFrame(ejes_mapeados, columns=["Front-Axis", "Final-Axis"])


        # ⭐⭐ AQUI AGREGAMOS TUS 8 PUNTOS A FUERZAS ⭐⭐
        forced_extra = pd.DataFrame([
            ["1100L[X]", "3125L[X]"],
            ["1100L[Y]", "3125L[Y]"],
            ["1100L[Z]", "3125L[Z]"],
            ["1100R[X]", "3125R[X]"],
            ["1100R[Y]", "3125R[Y]"],
            ["1100R[Z]", "3125R[Z]"]
        ], columns=["Front-Axis", "Final-Axis"])

        df_axes = pd.concat([df_axes, forced_extra], ignore_index=True).drop_duplicates()


        # ⭐⭐ FIN DEL BLOQUE NUEVO ⭐⭐


        final_ejes_validos = df_axes["Final-Axis"].unique()

        columnas_base = ["PSN", "JSN", "Fecha", "Hora"]

        df_final = df_final[[c for c in columnas_base if c in df_final.columns] + 
                            [e for e in final_ejes_validos if e in df_final.columns]]

        # --- FORZAR QUE LOS EJES EXTRA EXISTAN EN AMBOS DATAFRAMES ---
        for front_eje, final_eje in forced_extra.values:
            if front_eje not in df_frontal.columns:
                df_frontal[front_eje] = np.nan
            if final_eje not in df_final.columns:
                df_final[final_eje] = np.nan



        # --- Cálculo de correlaciones ---
        df_merge = pd.merge(
            df_frontal, df_final,
            on="PSN", suffixes=("_front", "_final")
        )

        correlacion_data = []

        for front_eje, final_eje in df_axes.values:

            # Intentar primero columnas sin sufijo
            col_front_1 = front_eje
            col_final_1 = final_eje

            # Intentar columnas con sufijo del merge
            col_front_2 = f"{front_eje}_front"
            col_final_2 = f"{final_eje}_final"

            # Determinar cuáles columnas existen en df_merge
            if col_front_1 in df_merge.columns and col_final_1 in df_merge.columns:
                col_front = col_front_1
                col_final = col_final_1
            elif col_front_2 in df_merge.columns and col_final_2 in df_merge.columns:
                col_front = col_front_2
                col_final = col_final_2
            else:
                # Si no existe ninguna forma, saltar eje
                continue

            tmp = df_merge[[col_front, col_final]].copy()

            tmp[col_front] = pd.to_numeric(tmp[col_front], errors="coerce")
            tmp[col_final] = pd.to_numeric(tmp[col_final], errors="coerce")

            tmp = tmp.dropna()

            if len(tmp) < 2:
                continue

            front_vals = tmp[col_front].values
            final_vals = tmp[col_final].values

            front_mean = np.mean(front_vals)
            final_mean = np.mean(final_vals)
            correlation = np.corrcoef(front_vals, final_vals)[0, 1]
            sigma6 = np.std(front_vals) * 6
            offset_calc = final_mean - front_mean

            correlacion_data.append([
                front_eje, final_eje,
                round(front_mean, 3),
                round(final_mean, 3),
                round(correlation, 3),
                round(sigma6, 3),
                round(offset_calc, 3)
            ])

 
 
        df_correlacion = pd.DataFrame(
            correlacion_data,
            columns=[
                "Front-Axis", "Final-Axis",
                "Front-Mean", "Final-Mean",
                "Correlation", "6Sigma", "Calculated-Offset"
            ]
        )
 
        def colorear_correlacion(val):
            if isinstance(val, (int, float)):
                if val >= 0.7:
                    return 'background-color: #47FF47; color: #000000; font-weight: 600;'
                elif val >= 0.69:
                    return 'background-color: #FFFD00; color: #000000; font-weight: 600;'
            return 'color: #FFFFFF;'
 
        def colorear_offset(val):
            if isinstance(val, (int, float)):
                if abs(val) > 1:
                    return 'background-color: #FF0000; color: #FFFFFF; font-weight: 600;'
                elif abs(val) > 0.5:
                    return 'background-color: #FFFD00; color: #000000; font-weight: 600;'
            return 'color: #FFFFFF;'
 
        df_correlacion_styled = (
            df_correlacion.style
            .applymap(colorear_correlacion, subset=["Correlation"])
            .applymap(colorear_offset, subset=["Calculated-Offset"])
            .set_table_styles([
                {'selector': 'th', 'props': [('background-color', '#2b2b2b'),
                                            ('color', '#FFFFFF'),
                                            ('font-weight', 'bold'),
                                            ('text-align', 'center'),
                                            ('padding', '8px')]},
                {'selector': 'td', 'props': [('background-color', '#1e1e1e'),
                                            ('color', '#FFFFFF'),
                                            ('text-align', 'center'),
                                            ('padding', '8px')]},
                {'selector': 'tbody tr:hover', 'props': [('background-color', '#333333')]},
                {'selector': 'table', 'props': [('border-radius', '10px'),
                                                ('overflow', 'hidden'),
                                                ('border', '1px solid #444')]}
            ])
        )
 
        st.subheader("📈 Correlación")
        st.dataframe(df_correlacion_styled, use_container_width=True)
 
        import openpyxl
        from openpyxl.styles import PatternFill, Font
 
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_frontal.to_excel(writer, index=False, sheet_name="Frontal")
            df_final.to_excel(writer, index=False, sheet_name="Final")
            df_match.to_excel(writer, index=False, sheet_name="Match_PSN")
            df_axes.to_excel(writer, index=False, sheet_name="Eje-Mapping")
            df_correlacion.to_excel(writer, index=False, sheet_name="Correlacion")

        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Correlacion"]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                if cell.value is not None:
                    if cell.value >= 0.7:
                        cell.fill = PatternFill(start_color="47FF47", end_color="47FF47", fill_type="solid")
                    elif cell.value >= 0.69:
                        cell.fill = PatternFill(start_color="FFFD00", end_color="FFFD00", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
            for cell in row:
                if cell.value is not None:
                    if abs(cell.value) > 1:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    elif abs(cell.value) > 0.5:
                        cell.fill = PatternFill(start_color="FFFD00", end_color="FFFD00", fill_type="solid")

        mean_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for col in [3, 4]:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                for cell in row:
                    if cell.value is not None:
                        cell.fill = mean_fill

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
 
        st.download_button(
            label="📥 Descargar Excel completo coloreado",
            data=excel_buffer,
            file_name="Mediciones_Percepton_Completo_Coloreado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
 
        # ---------------- XML ----------------
        def generar_xml_comparacion(df, station_name, model_name):
            import xml.etree.ElementTree as ET
           
            gauge = ET.Element("GAUGE")
            station = ET.SubElement(gauge, "STATION")
            ET.SubElement(station, "NAME").text = station_name
            model = ET.SubElement(station, "MODEL")
            ET.SubElement(model, "NAME").text = model_name
           
            df["Checkpoint"] = df["Front-Axis"].str.extract(r"(^\d+[LR])")
            df["Axis"] = df["Front-Axis"].str.extract(r"\[([XYZ])\]")
           
            for checkpoint_name, group in df.groupby("Checkpoint"):
                checkpoint = ET.SubElement(model, "CHECKPOINT")
                ET.SubElement(checkpoint, "NAME").text = checkpoint_name
               
                for axis in ["X", "Y", "Z"]:
                    axis_node = ET.SubElement(checkpoint, "AXIS")
                    ET.SubElement(axis_node, "NAME").text = axis
                    val = group.loc[group["Axis"]==axis, "Calculated-Offset"]
                    ET.SubElement(axis_node, "OFFSET").text = str(round(val.values[0],3)) if not val.empty else "0"
               
                axis_node = ET.SubElement(checkpoint, "AXIS")
                ET.SubElement(axis_node, "NAME").text = "Diameter"
                ET.SubElement(axis_node, "OFFSET").text = "0"
           
            xml_str = ET.tostring(gauge, encoding="utf-8", method="xml")
            return xml_str
 
        #xml_data = generar_xml_comparacion(df_correlacion)
        xml_data = generar_xml_comparacion(
            df_correlacion,
            station_name=station_name,
            model_name=model_name
        )
 
        st.download_button(
            label="📥 Descargar comparación en XML",
            data=xml_data,
            file_name="Comparacion_Percepton.xml",
            mime="application/xml"
        )
